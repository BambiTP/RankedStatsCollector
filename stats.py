import sys
import pandas as pd
import os
import re
import json
from openpyxl.utils import get_column_letter

# Check for input CSV argument
if len(sys.argv) < 2:
    print("Usage: python stats.py <combinedStatsMaster.csv>")
    sys.exit(1)
INPUT_CSV = sys.argv[1]

# Set root directory and load leaderboard
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
LEADERBOARD_FILE = os.path.join(ROOT_DIR, 'leaderboard.json')
if os.path.isfile(LEADERBOARD_FILE):
    with open(LEADERBOARD_FILE, 'r') as f:
        leaderboard = json.load(f)
else:
    leaderboard = {}
    print(f"Warning: {LEADERBOARD_FILE} not found; Skill column will be empty.")

# Define raw and derived statistics
raw_stats = [
    'Captures', 'Grabs', 'Hold', 'Drops', 'Pops', 'Returns', 'Tags', 'Prevent',
    'Pups', 'Pups Available', 'Block', 'Button', 'Support', 'Hold Against',
    'Long Holds', 'Flaccids', 'Handoffs', 'Good Handoffs', 'Captures off Handoffs',
    'Quick Returns', 'Key Returns', 'Returns in Base', 'NDPops', 'NRTags', 'KF', 'CD'
]

derived_stats = {
    'CD/Min': ('CD', 'Minutes'), 'Caps/Min': ('Captures', 'Minutes'),
    'Grabs/Min': ('Grabs', 'Minutes'), 'Hold/Min': ('Hold', 'Minutes'),
    'Drops/Min': ('Drops', 'Minutes'), 'Pops/Min': ('Pops', 'Minutes'),
    'Returns/Min': ('Returns', 'Minutes'), 'Tags/Min': ('Tags', 'Minutes'),
    'Prevent/Min': ('Prevent', 'Minutes'), 'Pups/Min': ('Pups', 'Minutes'),
    'Win %': ('Wins', 'Games'), 'Pup %': ('Pups', 'Pups Available'),
    'Score %': ('Captures', 'Grabs'), 'Flaccid %': ('Flaccids', 'Grabs'),
    'Chain %': ('Good Handoffs', 'Handoffs'), 'QR %': ('Quick Returns', 'Returns'),
    'RIB %': ('Returns in Base', 'Returns'), 'K/D': ('Tags', 'Pops'),
    'Hold/Grab': ('Hold', 'Grabs'), 'Prevent/Return': ('Prevent', 'Returns'),
    'Prevent/Hold Against': ('Prevent', 'Hold Against')
}

# Helper functions
def new_entry(name):
    return {
        'Name': name,
        'Games': 0,
        'RedGames': 0,
        'BlueGames': 0,
        'Wins': 0,
        'Losses': 0,
        'RedWins': 0,
        'BlueWins': 0,
        'Minutes': 0,
        'Totals': {stat: 0 for stat in raw_stats}
    }

def update_entry(entry, row, caps_for, caps_against, result, winner_color):
    tc = row.get('Team', '').strip().lower()
    entry['Games'] += 1
    if tc == 'red':
        entry['RedGames'] += 1
    elif tc == 'blue':
        entry['BlueGames'] += 1
    entry['Minutes'] += float(row.get('Minutes', 0))
    for stat in raw_stats:
        entry['Totals'][stat] += float(row.get(stat, 0))
    if result == 'win':
        entry['Wins'] += 1
        if tc == 'red':
            entry['RedWins'] += 1
        elif tc == 'blue':
            entry['BlueWins'] += 1
    else:
        entry['Losses'] += 1

def compute_derived(entry):
    dv = {}
    for label, (num_key, den_key) in derived_stats.items():
        num = entry['Totals'].get(num_key, entry.get(num_key, 0))
        den = entry['Totals'].get(den_key, entry.get(den_key, 0))
        dv[label] = num / den if den else 0
    minutes = entry['Minutes']
    for stat in raw_stats:
        total = entry['Totals'][stat]
        dv[f"{stat}/8Min"] = total / (minutes / 8) if minutes else 0
    return dv

def build_record(entry):
    rec = {
        'Player': entry['Name'],
        'Minutes': entry['Minutes'],
        'Games': entry['Games'],
        'Wins': entry['Wins'],
        'Losses': entry['Losses'],
        'Red Games': entry['RedGames'],
        'Blue Games': entry['BlueGames'],
        'Red Win %': entry['RedWins'] / entry['RedGames'] if entry['RedGames'] else 0,
        'Blue Win %': entry['BlueWins'] / entry['BlueGames'] if entry['BlueGames'] else 0
    }
    for stat in raw_stats:
        rec[stat] = entry['Totals'][stat]
    rec.update(compute_derived(entry))
    return rec

# Load and clean data
df = pd.read_csv(INPUT_CSV)
df.dropna(subset=['matchId', 'Player', 'Team'], inplace=True)

# Aggregate matches by matchId
matches = df.groupby('matchId')

# Process statistics
overall, per_map, map_results = {}, {}, {}
counted_map_games = set()

for match_id, match_df in matches:
    # Build true summed captures per team
    team_caps = (
        match_df
          .assign(team=match_df['Team'].str.strip().str.lower())
          .groupby('team')['Captures']
          .sum()
          .to_dict()
    )

    # Skip if not exactly two teams
    if len(team_caps) != 2:
        continue

    red_caps = team_caps.get('red', 0)
    blue_caps = team_caps.get('blue', 0)


    # Skip if highest individual minutes < 8 AND cap difference < 5 and not a tie
    max_minutes = match_df['Minutes'].max()
    if max_minutes < 8 and abs(red_caps - blue_caps) != 5 and abs(red_caps - blue_caps) > 0 :
        continue

    winner = 'red' if red_caps > blue_caps else 'blue'

    # Map-level results
    map_name = match_df['mapName'].dropna().iloc[0] if 'mapName' in match_df else None
    if map_name and (match_id, map_name) not in counted_map_games:
        mr = map_results.setdefault(map_name, {'Games': 0, 'RedWins': 0, 'BlueWins': 0})
        mr['Games'] += 1
        if winner == 'red':
            mr['RedWins'] += 1
        else:
            mr['BlueWins'] += 1
        counted_map_games.add((match_id, map_name))

    # Per-player and per-map accumulation
    for _, row in match_df.iterrows():
        key = row['Player'].strip().lower()
        overall.setdefault(key, new_entry(row['Player']))
        res = 'win' if row['Team'].strip().lower() == winner else 'loss'
        update_entry(overall[key], row,
                     caps_for=row['Captures'],
                     caps_against=team_caps[row['Team'].strip().lower()],
                     result=res,
                     winner_color=winner)

        if map_name:
            per_map.setdefault(map_name, {})
            per_map[map_name].setdefault(key, new_entry(row['Player']))
            update_entry(per_map[map_name][key], row,
                         caps_for=row['Captures'],
                         caps_against=team_caps[row['Team'].strip().lower()],
                         result=res,
                         winner_color=winner)
            team_key = f"{row['Team'].strip().lower()}_team"
            per_map[map_name].setdefault(team_key, new_entry(row['Team'].capitalize()))
            update_entry(per_map[map_name][team_key], row,
                         caps_for=row['Captures'],
                         caps_against=team_caps[row['Team'].strip().lower()],
                         result=res,
                         winner_color=winner)

# Build DataFrames
overall_df = pd.DataFrame([build_record(e) for e in overall.values()]) \
                 .sort_values(by='Minutes', ascending=False)
overall_df.insert(1, 'Skill', overall_df['Player'].map(lambda n: leaderboard.get(n, {}).get('skill', None)))

mr_df = pd.DataFrame([
    {
        'Map': m,
        'Games': v['Games'],
        'RedWins': v['RedWins'],
        'BlueWins': v['BlueWins'],
        'Red Win %': v['RedWins'] / v['Games'],
        'Blue Win %': v['BlueWins'] / v['Games']
    }
    for m, v in map_results.items()
])

per_csvs = []
for m, mp in per_map.items():
    dfm = pd.DataFrame([build_record(e) for e in mp.values()]) \
              .sort_values(by='Minutes', ascending=False)
    fn = f"stats_{m.replace(' ', '_').replace('/', '_')[:31]}.csv"
    per_csvs.append((fn, dfm))

# Create output directory
base = 'Stats'
os.makedirs(base, exist_ok=True)
existing = [
    int(match.group(1))
    for d in os.listdir(base)
    if (match := re.match(r'^Stats\((\d+)\)$', d)) and os.path.isdir(os.path.join(base, d))
]
idx = max(existing) + 1 if existing else 1
out = os.path.join(base, f"Stats({idx})")
os.makedirs(out)

# Save CSVs
overall_df.to_csv(os.path.join(out, 'players_stats_overall.csv'), index=False)
mr_df.to_csv(os.path.join(out, 'map_results.csv'), index=False)
for fn, dfm in per_csvs:
    dfm.to_csv(os.path.join(out, fn), index=False)

# Generate Excel workbook
excel_path = os.path.join(out, 'combined_stats.xlsx')
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    overall_df.to_excel(writer, sheet_name='OverallPlayers', index=False)
    mr_df.to_excel(writer, sheet_name='MapResults', index=False)
    for fn, dfm in per_csvs:
        raw = fn[:-4][:31]
        sheet = re.sub(r'[:\\\/?*\[\]]', '', raw)
        orig, i = sheet, 1
        while sheet in writer.sheets:
            sheet = (orig + f"_{i}")[:31]
            i += 1
        dfm.to_excel(writer, sheet_name=sheet, index=False)

    # Adjust column widths & formats
    for name, ws in writer.sheets.items():
        df_ref = (
            overall_df if name == 'OverallPlayers' else
            mr_df if name == 'MapResults' else
            next((df for fn, df in per_csvs
                  if name.startswith(re.sub(r'[:\\\/?*\[\]]', '', fn[:-4][:31]))), None)
        )
        if df_ref is None:
            continue

        pct_cols = [c for c in derived_stats if '%' in c] + ['Red Win %', 'Blue Win %']
        dec_cols = [c for c in derived_stats if '%' not in c] + ['Minutes'] \
                   + [c for c in df_ref.columns if c.endswith('/8Min')]

        for idx_col, col in enumerate(df_ref.columns, start=1):
            values = df_ref[col].fillna('')
            widths = [len(col)] + [
                len(f"{val:.2%}") if col in pct_cols else
                len(f"{val:.2f}") if col in dec_cols else
                len(str(val))
                for val in values
            ]
            ws.column_dimensions[get_column_letter(idx_col)].width = max(widths) + 2
            fmt = '0.00%' if col in pct_cols else '0.00' if col in dec_cols else None
            if fmt:
                for cell in ws[get_column_letter(idx_col)]:
                    cell.number_format = fmt

        ws.freeze_panes = 'B2'

print(f"Generated sorted stats and Excel workbook in {out}")
